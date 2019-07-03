# -*- coding: UTF-8 -*-
# Excel只有一个sheet

import os
from openpyxl import load_workbook
from config.excelConfig import SetExcel

proDir = os.path.split(os.path.realpath(__file__))[0]
excelPath = os.path.join(proDir, "../testDataFile/TestCase.xlsx")


class OperationExcel:
    def __init__(self):
        self.set_excel = SetExcel()
        self.open_excel = load_workbook(excelPath).active  # 打开Excel表格

    # 获取Excel表格的总行数
    def get_lines(self):
        line = self.open_excel.max_row
        return line

    # 通过单元格获取数据，例如：A2
    def from_cell_get_data(self, cell, row):
        """
        :param cell: 所在列A, B, ...
        :param row: 所在行1, 2, ...
        :return: 返回该单元格的数据
        """
        value = self.open_excel[cell + str(row)].value
        return value

    def from_coordinate_get_data(self, x, y):
        """通过单元格坐标获取数据，例如：(1, 2)
        :param x: 横坐标x
        :param y: 纵坐标y
        :return:返回该坐标(x, y)对应的数据
        """
        value = self.open_excel.cell(x, y).value
        return value

    def write_data(self, cell, row, write_value):
        """写入数据
        :param cell: 所在列A, B, ...
        :param row: 所在行1, 2, ...
        :param write_value: 写入的值
        :return:
        """
        wb = load_workbook(filename=excelPath)
        ws = wb.active
        ws[cell + str(row)] = write_value
        wb.save(filename=excelPath)


if __name__ == "__main__":
    s = OperationExcel()
    print(s.from_coordinate_get_data(1, 2))
