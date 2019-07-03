# -*- coding: UTF-8 -*-
# Excel有多个sheet

import os
from openpyxl import load_workbook
from config.excelConfig import SetExcel

proDir = os.path.split(os.path.realpath(__file__))[0]
excelPath = os.path.join(proDir, "../testDataFile/TestCase.xlsx")


class OperationExcel:
    def __init__(self):
        self.set_excel = SetExcel()
        self.open_excel = load_workbook(excelPath)  # 打开Excel表格

    def open_excel_sheet(self, sheet_name):
        """
        设置需要操作的sheet
        :param sheet_name:
        :return:
        """
        return self.open_excel[sheet_name]

    def get_lines(self, sheet_name):
        """
        获取Excel表格的总行数
        :param sheet_name:
        :return:
        """
        # line = self.open_excel.max_row
        open_sheet = self.open_excel_sheet(sheet_name)
        line = open_sheet.max_row
        return line

    def from_cell_get_data(self, sheet_name, cell, row):
        """
        通过单元格获取数据，例如：A2
        :param cell: 所在列A, B, ...
        :param row: 所在行1, 2, ...
        :return: 返回该单元格的数据
        """
        open_sheet = self.open_excel_sheet(sheet_name)
        value = open_sheet[cell + str(row)].value
        return value

    def from_coordinate_get_data(self, sheet_name, x, y):
        """
        通过单元格坐标获取数据，例如：(1, 2)
        :param row: 横坐标x
        :param column: 纵坐标y
        :return:返回该坐标(x, y)对应的数据
        """
        open_sheet = self.open_excel_sheet(sheet_name)
        value = open_sheet.cell(x, y).value
        return value

    def write_data(self, sheet_name, write_name, row, write_value):
        """
        写入数据
        :param write_name:
        :param row:
        :param write_value:
        :return:
        """
        wb = load_workbook(filename=excelPath)
        ws = wb[sheet_name]
        ws[write_name + str(row)] = write_value
        wb.save(filename=excelPath)


if __name__ == "__main__":
    s = OperationExcel()
    print(s.write_data('Case1', 'A', 3, 'aaa'))
